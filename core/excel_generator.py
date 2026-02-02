import os
import shutil
import tempfile
from copy import copy
import openpyxl
from openpyxl.worksheet.pagebreak import Break, PageBreak
from openpyxl.styles import Alignment
from datetime import datetime, timedelta

from config.settings import TEMPLATE_PATH, TEMP_DIR, EXCEL_START_ROW, EXCEL_TABLE_END_ROW, ROWS_PER_ITEM
from core.utils import format_address_for_excel, number_to_ringgit

class ExcelGenerator:
    def __init__(self):
        self.temp_filepath = None

    def generate_po_excel(self, po_data):
        print("Converting JSON to Excel...")
        
        gui_data = po_data.get('gui_data', {})
        
        # Create temporary file using context manager
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
            self.temp_filepath = temp_file.name
            
            try:
                workbook = openpyxl.load_workbook(TEMPLATE_PATH)
                sheet = workbook.active

                # Populate header with GUI data
                self._populate_header(sheet, gui_data)
                
                # Populate supplier information from quote
                self._populate_supplier_info(sheet, po_data, gui_data)
                
                # Populate items table
                total_cost, actual_rows_used = self._populate_items_table(sheet, po_data)
                
                # Add totals and formatting
                self._add_totals_and_formatting(sheet, total_cost, actual_rows_used, po_data)
                
                # Save to temporary file
                workbook.save(self.temp_filepath)
                print(f"✅ Successfully created temporary PO: {self.temp_filepath}")
                
                return self.temp_filepath

            except Exception as e:
                print(f"An unexpected error occurred: {e}")
                # Clean up temp file if error occurs
                if os.path.exists(self.temp_filepath):
                    os.unlink(self.temp_filepath)
                raise

    def _populate_header(self, sheet, gui_data):
        """Populate header information from GUI data"""
        sheet['E8'] = gui_data.get('po_number', '')
        sheet['H8'] = gui_data.get('po_issue_date', '')
        sheet['H9'] = gui_data.get('project_number', '')
        sheet['H10'] = gui_data.get('project_name', '')
        sheet['G18'] = f"Contact: {gui_data.get('purchaser_name', '')} ({gui_data.get('purchaser_phone', '')})"
        sheet['G20'] = f"PURCHASE ORDER NO.: {gui_data.get('po_number', '')}"

    def _populate_supplier_info(self, sheet, po_data, gui_data):
        """Populate supplier information from extracted PDF data"""
        sheet['B9'] = po_data.get('companyName')
        
        addr_line1, addr_line2 = format_address_for_excel(po_data.get('address', ''))
        sheet['B10'] = addr_line1
        sheet['B11'] = addr_line2
        
        sheet['C13'] = f": {po_data.get('pic', {}).get('name', '')}"
        sheet['C14'] = f": {po_data.get('pic', {}).get('phone', '')}"
        sheet['C15'] = f": {po_data.get('pic', {}).get('fax', '')}"
        sheet['C16'] = f": {po_data.get('pic', {}).get('email', '')}"
        sheet['A18'] = 'N/A'
        sheet['A21'] = po_data.get('terms', {}).get('payment')
        
        # Calculate delivery date
        delivery_weeks_str = str(po_data.get('terms', {}).get('deliveryWeeks', '0'))
        delivery_weeks_int = int(delivery_weeks_str) if delivery_weeks_str.isdigit() else 0
        if delivery_weeks_int > 0:
            po_issue_date = gui_data.get('po_issue_date', '')
            issue_date_obj = datetime.strptime(po_issue_date, "%d/%m/%Y")
            delivery_date = issue_date_obj + timedelta(weeks=delivery_weeks_int)
            sheet['H24'] = delivery_date.strftime("%d/%m/%Y")
        
        sheet['D29'] = f"With reference to your quotation {po_data.get('quotationNumber', '')}:"


    def _populate_items_table(self, sheet, po_data):
        """Populate items table and return total cost and actual rows used"""
        items_list = po_data.get('items', [])
        start_row = EXCEL_START_ROW
        
        # Calculate total rows needed and track item positions
        total_rows_needed = 0
        item_row_info = []  # Store (item, start_row, rows_needed)
        
        for item in items_list:
            description_rows = self._calculate_description_rows(item.get('description', ''))
            rows_needed = max(ROWS_PER_ITEM, description_rows + 1)  # At least 2 rows, more if description is long
            item_row_info.append((item, start_row + total_rows_needed, rows_needed))
            total_rows_needed += rows_needed
        
        # Check if we need to expand table
        available_rows = EXCEL_TABLE_END_ROW - EXCEL_START_ROW + 1
        if total_rows_needed > available_rows:
            self._expand_table_for_actual_rows(sheet, total_rows_needed, available_rows)
        
        # Populate items
        total_cost_calculated = 0
        for item, current_row, rows_needed in item_row_info:
            quantity = item.get('quantity') or 0
            unit_price = item.get('unitPrice') or 0
            line_total = quantity * unit_price
            total_cost_calculated += line_total

            # Get item index for numbering
            item_index = items_list.index(item) + 1
            
            sheet[f'A{current_row}'] = item_index
            sheet[f'B{current_row}'] = quantity
            sheet[f'C{current_row}'] = item.get('unit')
            self._populate_wrapped_description(sheet, item.get('description'), current_row)
            sheet[f'H{current_row}'] = unit_price
            sheet[f'I{current_row}'] = line_total

        return total_cost_calculated, total_rows_needed


    def _calculate_description_rows(self, description, max_chars_per_line=70):
        """Calculate how many rows a description needs based on character count"""
        if not description:
            return 1
        
        # Estimate characters per line (column D width)
        max_chars = max_chars_per_line
        words = description.split()
        lines_needed = 1
        current_line_length = 0
        
        for word in words:
            if current_line_length + len(word) + 1 <= max_chars:
                current_line_length += len(word) + 1
            else:
                lines_needed += 1
                current_line_length = len(word)
        
        return lines_needed

    def _populate_wrapped_description(self, sheet, description, start_row):
        """Populate description with proper wrapping across multiple rows"""
        if not description:
            sheet[f'D{start_row}'] = ''
            return
        
        max_chars = 70  # Adjust based on column width
        words = description.split()
        current_row = start_row
        current_line = []
        current_length = 0
        
        for word in words:
            if current_length + len(word) + 1 <= max_chars:
                current_line.append(word)
                current_length += len(word) + 1
            else:
                # Write current line
                sheet[f'D{current_row}'] = ' '.join(current_line)
                current_row += 1
                current_line = [word]
                current_length = len(word)
        
        # Write last line
        if current_line:
            sheet[f'D{current_row}'] = ' '.join(current_line)

    def _expand_items_table(self, sheet, num_items_to_add, available_item_slots, table_end_row):
        """Expand the items table if there are more items than available slots"""
        num_rows_to_insert = (num_items_to_add - available_item_slots) * ROWS_PER_ITEM
        insertion_point = table_end_row + 1

        merges_to_shift = []
        for merged_range in sheet.merged_cells.ranges:
            if merged_range.min_row >= insertion_point:
                merges_to_shift.append(merged_range)

        # Unmerge cells before inserting rows
        for merged_range in merges_to_shift:
            sheet.unmerge_cells(str(merged_range))

        # Insert blank rows
        sheet.insert_rows(insertion_point, amount=num_rows_to_insert)

        # Re-apply merges at new locations
        for merged_range in merges_to_shift:
            merged_range.shift(0, num_rows_to_insert)
            sheet.merge_cells(str(merged_range))

        # Copy styles to new rows (copy pattern from existing rows)
        for i in range(num_rows_to_insert):
            new_row_num = insertion_point + i
            # Copy style from the corresponding row in the original pattern
            pattern_row = EXCEL_START_ROW + (i % ROWS_PER_ITEM)
            self._copy_row_style(sheet, pattern_row, new_row_num)

    def _expand_table_for_actual_rows(self, sheet, total_rows_needed, available_rows):
        """Expand the items table based on actual row requirements"""
        num_rows_to_insert = total_rows_needed - available_rows
        insertion_point = EXCEL_TABLE_END_ROW + 1

        merges_to_shift = []
        for merged_range in sheet.merged_cells.ranges:
            if merged_range.min_row >= insertion_point:
                merges_to_shift.append(merged_range)

        # Unmerge cells before inserting rows
        for merged_range in merges_to_shift:
            sheet.unmerge_cells(str(merged_range))

        # Insert blank rows
        sheet.insert_rows(insertion_point, amount=num_rows_to_insert)

        # Re-apply merges at new locations
        for merged_range in merges_to_shift:
            merged_range.shift(0, num_rows_to_insert)
            sheet.merge_cells(str(merged_range))

        # Copy styles to new rows (copy pattern from existing rows)
        for i in range(num_rows_to_insert):
            new_row_num = insertion_point + i
            # Copy style from the corresponding row in the original pattern
            pattern_row = EXCEL_START_ROW + (i % ROWS_PER_ITEM)
            self._copy_row_style(sheet, pattern_row, new_row_num)

    def _copy_row_style(self, ws, source_row_num, dest_row_num):
        """Copy row style from source to destination"""
        source_row = ws[source_row_num]
        dest_row = ws[dest_row_num]

        ws.row_dimensions[dest_row_num].height = ws.row_dimensions[source_row_num].height

        for cell in source_row:
            new_cell = dest_row[cell.col_idx - 1]
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = cell.number_format
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)

    def _add_totals_and_formatting(self, sheet, total_cost, actual_rows_used, po_data):
        """Add totals, formatting, and signatures based on actual row usage"""
        items_list = po_data.get('items', [])
        num_items = len(items_list)
        print(f"Number of items: {num_items}")
        print(f"Actual rows used: {actual_rows_used}")
        
        # Calculate final table row based on actual rows used
        final_table_row = EXCEL_START_ROW + actual_rows_used - 1
        
        # Add total cost
        total_cost_row = final_table_row + 1
        sheet[f'I{total_cost_row}'] = f"=SUM(I{EXCEL_START_ROW}:I{final_table_row})"
        sheet['H12'] = f"=I{total_cost_row}"
        sheet.merge_cells('H12:I12')
        sheet['H12'].alignment = Alignment(horizontal='center')

        # Add total in words
        total_cost_in_words_row = final_table_row + 3
        sheet[f'E{total_cost_in_words_row}'] = number_to_ringgit(total_cost)
        
        # Add signatures
        name_rows = final_table_row + 10
        gui_data = po_data.get('gui_data', {})
        sheet[f'G{name_rows}'] = gui_data.get('purchaser_name', '')
        sheet[f'H{name_rows}'] = gui_data.get('director_manager', '')

        # Add page break
        sheet.row_breaks = PageBreak()
        page_break_row = final_table_row + 13
        sheet.row_breaks.append(Break(id=page_break_row))